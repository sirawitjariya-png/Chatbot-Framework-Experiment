"""Run both frameworks over the 100-question test set, 5 repeats each,
temperature=0, and save raw results for RAGAS scoring + paired stats.

Usage:
    python -m experiments.run_experiment
    python -m experiments.run_experiment --framework langgraph
    python -m experiments.run_experiment --limit 5      # smoke test

Output: experiments/results/raw_runs.csv
  ONE ROW per (question_id, framework). Every one of the 5 repeats gets its
  own numbered column (answer_1..answer_5, cost_thb_1..cost_thb_5,
  latency_s_1..latency_s_5, ...) so every individual run is visible side by
  side, plus cost_thb_avg/cost_thb_sd and latency_s_avg/latency_s_sd columns
  that summarize the 5 repeats (sd = sample standard deviation, i.e. how
  much cost/latency actually varied run to run at temperature=0). This is
  the input to experiments/ragas_eval.py.
"""
import argparse
import csv
import logging
import statistics
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

from shared.config import (
    N_REPEATS_PER_QUESTION, PRICE_INPUT_THB_PER_1M, PRICE_OUTPUT_THB_PER_1M,
    PRICE_CACHED_INPUT_THB_PER_1M, REPO_ROOT,
)

RESULTS_DIR = REPO_ROOT / "experiments" / "results"
RESULTS_DIR.mkdir(parents=True, exist_ok=True)
LOG_FILE = RESULTS_DIR / "run_experiment.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[logging.StreamHandler(), logging.FileHandler(LOG_FILE, mode="w", encoding="utf-8")],
)
log = logging.getLogger(__name__)

QUESTIONS_XLSX = REPO_ROOT / "tests" / "hospital_chatbot_test_questions.xlsx"
OUTPUT_CSV = RESULTS_DIR / "raw_runs.csv"

FRAMEWORKS = {
    "langgraph": "frameworks.langgraph_impl.graph",
    "skillsmd": "frameworks.skillsmd_impl.agent",
}

# Per-repeat fields: one column per repeat (1..N_REPEATS_PER_QUESTION) for each.
PER_REPEAT_FIELDS = [
    "answer", "context", "route", "files",
    "in_tok", "out_tok", "cached_tok", "latency_s", "n_llm_calls", "cost_thb", "error",
]
# Of those, these also get an "_avg" column across the N repeats.
AVERAGED_FIELDS = ["cost_thb", "latency_s"]


def load_questions(limit: int | None = None, question_ids: list[int] | None = None) -> list[dict]:
    wb = openpyxl.load_workbook(QUESTIONS_XLSX, data_only=True)
    ws = wb["Test Questions"]
    hdr = [c.value for c in ws[1]]

    def col(name_substr):
        for i, h in enumerate(hdr):
            if h and name_substr.lower() in str(h).lower():
                return i
        return None

    ci_no, ci_cat, ci_q = col("No."), col("Category"), col("Question")
    wanted = set(question_ids) if question_ids else None
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[ci_no] is None:
            continue
        if wanted is not None and row[ci_no] not in wanted:
            continue
        rows.append({"question_id": row[ci_no], "category": row[ci_cat], "question": row[ci_q]})
        if limit and len(rows) >= limit:
            break
    return rows


def estimate_cost_thb(in_tok: int, out_tok: int, cached_tok: int) -> float:
    billable_in = max(in_tok - cached_tok, 0)
    cost = (
        billable_in / 1_000_000 * PRICE_INPUT_THB_PER_1M
        + cached_tok / 1_000_000 * PRICE_CACHED_INPUT_THB_PER_1M
        + out_tok / 1_000_000 * PRICE_OUTPUT_THB_PER_1M
    )
    return round(cost, 6)


def run_one(ask_fn, question: str) -> dict:
    t0 = time.perf_counter()
    try:
        result = ask_fn(question, history=[])
    except Exception as e:
        log.error("ask() failed: %s", e)
        return {"answer": "", "context": "", "route": None, "files": [], "in_tok": 0, "out_tok": 0,
                "cached_tok": 0, "latency_s": round(time.perf_counter() - t0, 4),
                "n_llm_calls": 0, "cost_thb": 0.0, "error": str(e)}

    metrics = result.get("metrics", [])
    total_in = sum(m["in_tok"] for m in metrics)
    total_out = sum(m["out_tok"] for m in metrics)
    total_cached = sum(m.get("cached_tok", 0) for m in metrics)
    total_latency = time.perf_counter() - t0  # wall-clock, not summed per-call latency
    return {
        "answer": result.get("answer", ""),
        "context": result.get("context", ""),
        "route": result.get("route"),
        "files": result.get("files", []),
        "in_tok": total_in,
        "out_tok": total_out,
        "cached_tok": total_cached,
        "latency_s": round(total_latency, 4),
        "n_llm_calls": len(metrics),
        "cost_thb": estimate_cost_thb(total_in, total_out, total_cached),
        "error": "",
    }


def build_fieldnames(n_repeats: int) -> list[str]:
    fields = ["question_id", "category", "question", "framework"]
    for base in PER_REPEAT_FIELDS:
        for rep in range(1, n_repeats + 1):
            fields.append(f"{base}_{rep}")
        if base in AVERAGED_FIELDS:
            fields.append(f"{base}_avg")
            fields.append(f"{base}_sd")
    return fields


def mean_sd(values: list[float]) -> tuple[float, float]:
    """Sample standard deviation (ddof=1); 0.0 if fewer than 2 values."""
    mean = sum(values) / len(values)
    sd = statistics.stdev(values) if len(values) > 1 else 0.0
    return mean, sd


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--framework", choices=list(FRAMEWORKS.keys()), default=None,
                         help="Run only one framework (default: both)")
    parser.add_argument("--limit", type=int, default=None, help="Only run first N questions (smoke test)")
    parser.add_argument("--question-ids", type=str, default=None,
                         help="Comma-separated question numbers to run, e.g. --question-ids 1,15,45,70,100 "
                              "(a pilot subset instead of the full test set)")
    parser.add_argument("--repeats", type=int, default=N_REPEATS_PER_QUESTION)
    parser.add_argument("--append", action="store_true", help="Append to existing CSV instead of overwriting")
    args = parser.parse_args()

    question_ids = [int(x) for x in args.question_ids.split(",")] if args.question_ids else None
    questions = load_questions(limit=args.limit, question_ids=question_ids)
    log.info("Loaded %d questions", len(questions))

    frameworks_to_run = [args.framework] if args.framework else list(FRAMEWORKS.keys())
    fieldnames = build_fieldnames(args.repeats)

    OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    mode = "a" if (args.append and OUTPUT_CSV.exists()) else "w"
    write_header = mode == "w"

    with open(OUTPUT_CSV, mode, newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        if write_header:
            writer.writeheader()

        for fw_name in frameworks_to_run:
            module_path = FRAMEWORKS[fw_name]
            module = __import__(module_path, fromlist=["ask"])
            ask_fn = module.ask
            log.info("=== Framework: %s ===", fw_name)

            for q in questions:
                row = {"question_id": q["question_id"], "category": q["category"],
                       "question": q["question"], "framework": fw_name}
                cost_vals, latency_vals = [], []

                for rep in range(1, args.repeats + 1):
                    log.info("[%s] Q%s rep %d/%d: %s", fw_name, q["question_id"], rep,
                              args.repeats, str(q["question"])[:50])
                    r = run_one(ask_fn, q["question"])
                    for base in PER_REPEAT_FIELDS:
                        row[f"{base}_{rep}"] = r[base]
                    cost_vals.append(r["cost_thb"])
                    latency_vals.append(r["latency_s"])

                cost_mean, cost_sd = mean_sd(cost_vals)
                latency_mean, latency_sd = mean_sd(latency_vals)
                row["cost_thb_avg"] = round(cost_mean, 6)
                row["cost_thb_sd"] = round(cost_sd, 6)
                row["latency_s_avg"] = round(latency_mean, 4)
                row["latency_s_sd"] = round(latency_sd, 4)

                writer.writerow(row)
                f.flush()

    log.info("Done. Results written to %s", OUTPUT_CSV)
    log.info("Log written to %s", LOG_FILE)


if __name__ == "__main__":
    main()
